from pathlib import Path

import openpyxl
from rdflib import URIRef, Namespace, Graph, RDF, Dataset, DCAT, RDFS, Literal

OLIS = Namespace("https://olis.dev/")
THEME_GRAPH = URIRef("https://themes")

filename = "SA Knowledge Base Library construction v6.xlsx"
sheetname = "categorisation of Vocabs"

file = Path(__file__).parent.parent.parent / "data/categories/" / filename
workbook = openpyxl.load_workbook(file)

sheet = workbook[sheetname]

# Dictionary to store the IRI as keys and SeaDataNet keyword types and URIs as values
iri_dict = {}

# Iterate through the rows in the specified columns
for row in sheet.iter_rows(
    min_row=2, max_row=sheet.max_row, min_col=2, max_col=3, values_only=True
):
    iri_value, sea_data_net_value = row
    if iri_value:
        real_g_uri = URIRef(iri_value)
        # Check if both columns have values
        if real_g_uri and sea_data_net_value:
            # If the IRI value already exists in the dictionary, append the new value
            if real_g_uri in iri_dict:
                iri_dict[real_g_uri].append(sea_data_net_value)
            else:
                # If not, create a new list with the SeaDataNet keyword type and URIs as the initial value
                iri_dict[real_g_uri] = [sea_data_net_value]

string_iri_map = {
    "parameter": URIRef("http://vocab.nerc.ac.uk/collection/L19/current/SDNKG03/"),
    "instrument": URIRef("http://vocab.nerc.ac.uk/collection/L19/current/SDNKG01/"),
    "platform": URIRef("http://vocab.nerc.ac.uk/collection/L19/current/SDNKG04/"),
}

# olis_system_graph = Graph(identifier=OLIS.System)
# theme_graph = Graph(identifier=THEME_GRAPH)
ds = Dataset(default_union=True)
# create virtual graphs for the different categories
for label, virt_g in string_iri_map.items():
    ds.add((virt_g, RDF.type, OLIS.VirtualGraph, OLIS.System))
    ds.add((virt_g, RDFS.label, Literal(label.capitalize()), THEME_GRAPH))

for real_g_uri, category_list in iri_dict.items():
    split_category_list = category_list[0].split("; ")
    for category_str in split_category_list:
        ds.add((string_iri_map[category_str], OLIS.isAliasFor, real_g_uri, OLIS.System))
        ds.add((real_g_uri, DCAT.theme, string_iri_map[category_str], THEME_GRAPH))

ds.serialize(
    Path(__file__).parent.parent.parent / "output/graph_categories/graph_categories.nq",
    format="nquads",
)

# for s,p,o,g in ds:
#     print(s,p,o,g)
